import json
import os
import sys
import openpyxl 

cur_dir = os.path.dirname(__file__)
helper_dir = os.path.join(cur_dir, "..", "tools")
sys.path.append(helper_dir)
from config import group_accounts_board_id, measure_tracking_board_id, mtb_board_id
from tools import query_helper, write_data 

group_list = query_helper(
    "query($board_id:Int!, $column_id:String!, $group_name: String!) {boards(ids:[$board_id]){groups(ids:[$group_name]){items{id column_values(ids:[$column_id]){text}}}}}",
    {"board_id": group_accounts_board_id, "column_id": "text9", "group_name" : "1660067352_group_names"}, 
)["data"]["boards"][0]["groups"][0]["items"]

template_measure_data = query_helper(
    "query($board_id:Int!) {boards(ids:[$board_id]){items{name}}}",
    {"board_id": mtb_board_id}, 
)["data"]["boards"][0]["items"]

template_measure_list=[]
for measure in template_measure_data:
    template_measure_list.append(measure["name"])

new_dir = cur_dir +"\\data\\new_mtb.xlsx" 
new_mtb_wb = openpyxl.Workbook()
new_mtb_sheet = new_mtb_wb.active
name_row = 1

new_dir2 = cur_dir + "\\data\\measure_data.xlsx"
measure_wb = openpyxl.Workbook()
measure_sheet = measure_wb.active

group_row = 1

measure_sheet.cell(row=group_row, column=1,value="Group Name")
group_row+=1


for group_accounts in group_list:
    group_name = group_accounts["column_values"][0]["text"]
    measure_data = query_helper(
        "query($board_id:Int!, $column_id:String!, $column_val:String!) {items_by_column_values (board_id: $board_id, column_id: $column_id, column_value: $column_val){name id}}",
        {"board_id": measure_tracking_board_id, "column_id": "dup__of_group_name4", "column_val": group_name},
    )["data"]["items_by_column_values"]

    measure_tuple_list=[]
    measure_list=[]
    for measure in measure_data:
        if "QPP 326" in measure["name"]:
            measure_tuple : tuple[str | int]= (measure["name"].split(";")[0],int(measure["id"]))
            measure_tuple_list.append(measure_tuple)
            measure_list.append(measure["name"].split(";")[0])
        else:
            measure_tuple : tuple[str | int]= (measure["name"].replace("\n", " "),int(measure["id"]))
            measure_tuple_list.append(measure_tuple)
            measure_list.append(measure["name"].replace("\n", " "))

    retired_measures=[]
    for measure_tuple in measure_tuple_list:
        for measure_name in measure_tuple:
            if isinstance(measure_name,int) is False:
                if measure_name not in template_measure_list:
                    retired_measures.append(measure_tuple)

    missing_measures = [template_measure for template_measure in template_measure_list if template_measure not in measure_list]

    if len(missing_measures)==48:
        new_mtb_sheet.cell(row=name_row, column=1,value=group_name)
        name_row+=1
   
    measure_sheet.cell(row=group_row,column=1,value=group_name)
    group_row+=1
    
    measure_sheet.cell(row=group_row, column=2,value="Added Measures")
    group_row+=1
    
    for measure in missing_measures:
        measure_sheet.cell(row=group_row,column=2,value=measure)
        group_row+=1
        
    measure_sheet.cell(row=group_row,column=2,value="Retired Measures")
    group_row+=1
    
    for measure_tuple in retired_measures:
        for measure_name in measure_tuple:
            if isinstance(measure_name,int) is False:
                measure_sheet.cell(row=group_row,column=2,value=measure_name)
                group_row+=1

    for measure_tuple in retired_measures:
        for measure_id in measure_tuple:
            if isinstance(measure_id,int) is True:
                col_id="status"
                query_helper(
                    "mutation ($board_id:Int!,$item_id:Int!, $col_val: JSON!) { change_multiple_column_values(item_id: $item_id, board_id: $board_id, column_values: $col_val){id}}",
                    {"board_id":measure_tracking_board_id,
                    "item_id":measure_id,
                    "col_val": json.dumps({col_id: {"label": "Retired"}})})

    for measure in missing_measures:
        col_id = "dup__of_group_name4"
        query_helper("mutation ($board_id:Int!,$item_name:String!, $col_val: JSON!) {create_item(item_name: $item_name, board_id: $board_id, column_values: $col_val){id}}",
            {"board_id":measure_tracking_board_id,
            "item_name": measure,
            "col_val": json.dumps({col_id: group_name})
            },True)



new_mtb_wb.save(new_dir)
measure_wb.save(new_dir2)
    
   






