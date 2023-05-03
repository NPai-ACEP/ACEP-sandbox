import json
import os
import sys
import openpyxl

cur_dir = os.path.dirname(__file__)
helper_dir = os.path.join(cur_dir, "..", "tools")
sys.path.append(helper_dir)
from config import measure_tracking_board_id
from tools import query_helper

group_accounts_overview_ingest = query_helper(
    "query($board_id:Int!){boards(ids:[$board_id]){groups {title items{id}}}}",
    {
        "board_id": ed_mgmt_board_id,
    },
)
group_accounts_overview_list = group_accounts_overview_ingest["data"]["boards"][0][
    "groups"
]

group_accounts_ingest = query_helper(
    "query($board_id:Int!, $col_id:String!,$col_id1:String!,$col_id2:String!){boards(ids:[$board_id]){ items{name id column_values(ids:[$col_id,$col_id1,$col_id2]){text value}}}}",
    {
        "board_id": ed_mgmt_board_id,
        "col_id": "text7",
        "col_id1": "lookup",
        "col_id2": "connect_boards44",
    },
)
group_accounts_list = group_accounts_ingest["data"]["boards"][0]["items"]

ed_data_ingest = query_helper(
    "query($board_id:Int!, $col_id:String!){boards(ids:[$board_id]){ items{name id column_values(ids:[$col_id]){text}}}}",
    {
        "board_id": ed_mgmt_board_id,
        "col_id": "text4",
    },
)
ed_data_list = ed_data_ingest["data"]["boards"][0]["items"]


ed_data_wb = openpyxl.Workbook()
ed_data_sheet = ed_data_wb.active
status_row_num = 1
group_row_num = 0
ed_row_num = 0
for status in group_accounts_overview_list:
    ed_data_sheet.cell(row=status_row_num, column=1, value=("Group Status"))

    status_row_num += 1
    ed_data_sheet.cell(row=status_row_num, column=1, value=(status["title"]))

    status_row_num += 1
    group_row_num = status_row_num
    for overview in status["items"]:
        for group in group_accounts_list:
            if group_row_num == status_row_num and overview["id"] == group["id"]:
                status_row_num = len(status["items"]) + status_row_num + 1
                ed_data_sheet.cell(row=group_row_num, column=2, value=("Group ID"))
                ed_data_sheet.cell(row=group_row_num, column=3, value=("Group Name"))
                ed_data_sheet.cell(row=group_row_num, column=4, value=("Group Status"))
                group_row_num += 1
                ed_data_sheet.cell(
                    row=group_row_num,
                    column=2,
                    value=(int(group["column_values"][0]["text"])),
                )
                ed_data_sheet.cell(row=group_row_num, column=3, value=(group["name"]))
                ed_data_sheet.cell(
                    row=group_row_num,
                    column=4,
                    value=(group["column_values"][1]["text"]),
                )
                group_row_num += 1
                ed_row_num = group_row_num
                ed_grp_ids = json.loads(group["column_values"][2]["value"])[
                    "linkedPulseIds"
                ]
                for ed in ed_data_list:
                    for ed_grp_id in ed_grp_ids:
                        if ed_row_num == group_row_num and int(
                            ed_grp_id["linkedPulseId"]
                        ) == int(ed["id"]):
                            group_row_num = len(ed_grp_ids) + group_row_num + 1
                            status_row_num = len(status["items"]) + group_row_num + 1
                            ed_data_sheet.cell(
                                row=ed_row_num, column=3, value=("ED ID")
                            )
                            ed_data_sheet.cell(
                                row=ed_row_num, column=4, value=("ED Name")
                            )
                            ed_data_sheet.cell(
                                row=ed_row_num, column=5, value=("ED Status")
                            )
                            ed_row_num += 1
                            ed_data_sheet.cell(
                                row=ed_row_num,
                                column=3,
                                value=(int(ed["column_values"][0]["text"])),
                            )
                            ed_data_sheet.cell(
                                row=ed_row_num, column=4, value=(ed["name"])
                            )
                            ed_data_sheet.cell(
                                row=ed_row_num,
                                column=5,
                                value=(group["column_values"][1]["text"]),
                            )
                            ed_row_num += 1
                        elif ed_row_num < group_row_num and int(
                            ed_grp_id["linkedPulseId"]
                        ) == int(ed["id"]):
                            ed_data_sheet.cell(
                                row=ed_row_num,
                                column=3,
                                value=(int(ed["column_values"][0]["text"])),
                            )
                            ed_data_sheet.cell(
                                row=ed_row_num, column=4, value=(ed["name"])
                            )
                            ed_data_sheet.cell(
                                row=ed_row_num,
                                column=5,
                                value=(group["column_values"][1]["text"]),
                            )
                            ed_row_num += 1
                # group_row_num = len(group[""])
            elif group_row_num < status_row_num and overview["id"] == group["id"]:
                ed_data_sheet.cell(
                    row=group_row_num,
                    column=2,
                    value=(int(group["column_values"][0]["text"])),
                )
                ed_data_sheet.cell(row=group_row_num, column=3, value=(group["name"]))
                ed_data_sheet.cell(
                    row=group_row_num,
                    column=4,
                    value=(group["column_values"][1]["text"]),
                )
                group_row_num += 1
                ed_row_num = group_row_num
                ed_grp_ids = json.loads(group["column_values"][2]["value"])[
                    "linkedPulseIds"
                ]
                ed_row_num = group_row_num
                for ed in ed_data_list:
                    for ed_grp_id in ed_grp_ids:
                        if ed_row_num == group_row_num and int(
                            ed_grp_id["linkedPulseId"]
                        ) == int(ed["id"]):
                            group_row_num = len(ed_grp_ids) + group_row_num + 1
                            status_row_num = len(status["items"]) + group_row_num + 1
                            ed_data_sheet.cell(
                                row=ed_row_num, column=3, value=("ED ID")
                            )
                            ed_data_sheet.cell(
                                row=ed_row_num, column=4, value=("ED Name")
                            )
                            ed_data_sheet.cell(
                                row=ed_row_num, column=5, value=("ED Status")
                            )
                            ed_row_num += 1
                            ed_data_sheet.cell(
                                row=ed_row_num,
                                column=3,
                                value=(int(ed["column_values"][0]["text"])),
                            )
                            ed_data_sheet.cell(
                                row=ed_row_num, column=4, value=(ed["name"])
                            )
                            ed_data_sheet.cell(
                                row=ed_row_num,
                                column=5,
                                value=(group["column_values"][1]["text"]),
                            )
                            ed_row_num += 1
                        elif ed_row_num < group_row_num and int(
                            ed_grp_id["linkedPulseId"]
                        ) == int(ed["id"]):
                            ed_data_sheet.cell(
                                row=ed_row_num,
                                column=3,
                                value=(int(ed["column_values"][0]["text"])),
                            )
                            ed_data_sheet.cell(
                                row=ed_row_num, column=4, value=(ed["name"])
                            )
                            ed_data_sheet.cell(
                                row=ed_row_num,
                                column=5,
                                value=(group["column_values"][1]["text"]),
                            )
                            ed_row_num += 1
ed_data_wb.save("/home/npai/ACEP-sandbox/data_extract/data/ed_status_data.xlsx")
